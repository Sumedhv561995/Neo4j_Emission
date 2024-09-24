import json
import openpyxl
from py2neo import Graph
import boto3
from io import BytesIO

# Load the configuration from config.json
with open('Schema.json') as f:
    config = json.load(f)

# Neo4j connection settings
neo4j_uri = config['neo4j']['uri']
username = config['neo4j']['username']
password = config['neo4j']['password']

# Create an S3 client
s3 = boto3.client('s3')

# Define the bucket name and the file names (keys)
bucket_name = config['s3']['bucket_name']  # Fetch the bucket name from config file
excel_file = config['s3']['sap_hierarchy_file']  # Fetch the file name for hierarchy from config
additional_excel_file = config['s3']['operating_parameters_file']  # Fetch the file name for operating parameters


# Read the file content from S3 as binary
def read_file_from_s3(bucket, key):
    try:
        obj = s3.get_object(Bucket=bucket, Key=key)
        return obj['Body'].read()  # Return binary data
    except Exception as e:
        print(f"Error occurred: {e}")
        return None


# Load the Excel files from S3
file_content = read_file_from_s3(bucket_name, excel_file)
file_content1 = read_file_from_s3(bucket_name, additional_excel_file)

# Ensure the file content is loaded
if file_content and file_content1:
    # Load the Excel workbooks from binary content
    wb_hierarchy = openpyxl.load_workbook(BytesIO(file_content))
    wb_operating_params = openpyxl.load_workbook(BytesIO(file_content1))

    # Get sheet names
    sheet_names = wb_hierarchy.sheetnames
    print("Sheet names from SAP Hierarchy:", sheet_names)

    operating_sheet_names = wb_operating_params.sheetnames
    print("Sheet names from Operating Parameters:", operating_sheet_names)

    # Connect to Neo4j
    graph = Graph(neo4j_uri, auth=(username, password))

    # Define the relationships list from the config file
    relationships = config['relationships']

    # Create the Northstar_Platform node
    northstar_platform_name = config['northstar_platform']['name']
    try:
        graph.run("MERGE (n:Northstar_Platform {name: $name})", name=northstar_platform_name)
        print("Northstar_Platform node ensured.")
    except Exception as e:
        print(f"Error creating Northstar_Platform node: {e}")


    # Function to create a Cypher query for merging a node
    def create_node_query(sheet, header_row, row_data):
        labels = sheet.title  # Use the sheet name as the node label
        properties = {header_row[i]: row_data[i] for i in range(len(header_row)) if row_data[i]}
        return {"query": f"MERGE (n:{labels} {{ {', '.join([f'{key}: $props.{key}' for key in properties.keys()])} }})",
                "parameters": {"props": properties}}


    # Function to link IFLOT nodes to the Northstar_Platform using the Description column
    def link_iflot_to_northstar(description_value):
        return {
            "query": f"""
                MATCH (fl:IFLOT {{Description: $description_value}})
                MATCH (p:Northstar_Platform {{name: $platform_name}})
                MERGE (p)-[:HAS]->(fl)
            """,
            "parameters": {"description_value": description_value, "platform_name": northstar_platform_name}
        }


    # Function to create and link parameters from the operating parameters file
    def create_parameter_nodes(sheet, header_row, row_data, equipment_name):
        # Skip the timestamp column and create child nodes for the parameters
        parameter_queries = []
        timestamp = row_data[1]  # Assume timestamp is the first column

        # Cleaning up the equipment_name to avoid mismatches due to leading/trailing spaces
        equipment_name_cleaned = equipment_name.strip()

        # Ensure the equipment node exists in Neo4j by matching it with the EQUI.Description
        equipment_query = {
            "query": """
                    MATCH (e:EQUI {Description: $equipment_name})
                    RETURN e
                """,
            "parameters": {
                "equipment_name": equipment_name_cleaned
            }
        }

        try:
            # Ensure equipment node exists
            graph.run(equipment_query["query"], **equipment_query["parameters"])
            print(f"Equipment node '{equipment_name_cleaned}' ensured.")
        except Exception as e:
            print(f"Error ensuring equipment node: {e}")

        # Create parameter nodes and link them to the equipment node
        for i in range(2, len(header_row)):
            parameter_name = header_row[i]
            parameter_value = row_data[i]
            if parameter_value is not None:
                parameter_query = {
                    "query": f"""
                            MATCH (e:EQUI {{Description: $equipment_name}})
                            MERGE (e)-[:HAS_PARAMETER]->(p:Parameter {{name: $parameter_name}})
                            MERGE (p)-[:HAS_VALUE]->(v:Value {{value: $parameter_value, timestamp: $timestamp}})
                        """,
                    "parameters": {
                        "equipment_name": equipment_name_cleaned,
                        "parameter_name": parameter_name,
                        "parameter_value": parameter_value,
                        "timestamp": timestamp
                    }
                }
                parameter_queries.append(parameter_query)
        return parameter_queries


    # Step 1: Process the SAP Hierarchy file and create nodes
    for sheet_name in wb_hierarchy.sheetnames:
        sheet = wb_hierarchy[sheet_name]
        header_row = [cell.value for cell in sheet[1]]
        node_data = []

        # Collect data to create nodes
        for row in sheet.iter_rows(min_row=2, values_only=True):
            node_query = create_node_query(sheet, header_row, row)
            node_data.append(node_query)

        # Execute node creation queries in batch
        for node_query in node_data:
            try:
                graph.run(node_query["query"], **node_query["parameters"])
                print(f"Node for sheet '{sheet_name}' created/merged.")
            except Exception as e:
                print(f"Error creating node in sheet '{sheet_name}': {e}")

            # Step 1.1: If sheet is IFLOT, link nodes to the Northstar_Platform using Description
            if sheet_name == "IFLOT":
                try:
                    description_index = header_row.index("Description")  # Ensure Description column exists
                except ValueError:
                    print(f"Error: 'Description' column not found in sheet '{sheet_name}'.")
                    continue

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    description_value = row[description_index]
                    if description_value:
                        try:
                            link_query = link_iflot_to_northstar(description_value)
                            graph.run(link_query["query"], **link_query["parameters"])
                            print(f"IFLOT node with description '{description_value}' linked to Northstar_Platform.")
                        except Exception as e:
                            print(
                                f"Error linking IFLOT node with description '{description_value}' to Northstar_Platform: {e}")


    # Function to link parameter sheets to existing EQUI nodes dynamically
    def link_parameters_to_existing_equi(sheet, header_row, row_data):
        # Assuming the sheet name is the equipment name
        equipment_name = sheet.title.strip()
        timestamp = row_data[1]  # Assume the timestamp is in the first column

        # Ensure the equipment node already exists in Neo4j, if not, skip
        equipment_query = {
            "query": """
                MATCH (e:EQUI {Description: $equipment_name})
                RETURN e
            """,
            "parameters": {
                "equipment_name": equipment_name
            }
        }

        equipment_exists = graph.run(equipment_query["query"], **equipment_query["parameters"]).data()

        # If the equipment node exists, proceed to create parameters and relationships
        if equipment_exists:
            print(
                f"Found existing equipment node for '{equipment_name}', proceeding to create parameter relationships.")

            # Iterate through each parameter in the row (skipping the timestamp)
            for i in range(2, len(header_row)):
                parameter_name = header_row[i]
                parameter_value = row_data[i]

                if parameter_value is not None:
                    parameter_query = {
                        "query": """
                            MATCH (e:EQUI {Description: $equipment_name})
                            MERGE (e)-[:HAS_PARAMETER]->(p:Parameter {name: $parameter_name})
                            MERGE (p)-[:HAS_VALUE]->(v:Value {value: $parameter_value, timestamp: $timestamp})
                        """,
                        "parameters": {
                            "equipment_name": equipment_name,
                            "parameter_name": parameter_name,
                            "parameter_value": parameter_value,
                            "timestamp": timestamp
                        }
                    }

                    # Execute the query to link the parameters to the existing EQUI node
                    try:
                        graph.run(parameter_query["query"], **parameter_query["parameters"])
                        print(f"Parameter '{parameter_name}' for equipment '{equipment_name}' created/linked.")
                    except Exception as e:
                        print(
                            f"Error creating parameter node for '{parameter_name}' on equipment '{equipment_name}': {e}")
        else:
            print(
                f"Equipment '{equipment_name}' does not exist in Neo4j. Skipping parameter creation for this equipment.")


    # Step 2: Process the Operating Parameters file and create relationships
    for sheet_name in wb_operating_params.sheetnames:
        sheet = wb_operating_params[sheet_name]
        header_row = [cell.value for cell in sheet[1]]
        print(sheet)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            link_parameters_to_existing_equi(sheet, header_row, row)
            equipment_name = sheet_name  # Assuming sheet_name is the equipment name
            equipment_name_cleaned = equipment_name.strip()

            # Create the parameter nodes and link them to the equipment node
            parameter_queries = create_parameter_nodes(sheet, header_row, row, equipment_name_cleaned)

            # Execute parameter node creation queries
            for parameter_query in parameter_queries:
                try:
                    graph.run(parameter_query["query"], **parameter_query["parameters"])
                    print(f"Parameters for equipment '{equipment_name_cleaned}' created/linked.")
                except Exception as e:
                    print(f"Error creating parameter node: {e}")

    # Step 3: Create additional relationships (if any) from config
    for relationship in relationships:
        sheet_name = relationship["sheet_name"]
        join_column = relationship["join_column"]
        target_sheet_name = relationship["target_sheet_name"]
        target_join_column = relationship["target_join_column"]
        relationship_type = relationship["relationship_type"]

        print(f"\nProcessing relationship: {sheet_name} -> {target_sheet_name} as {relationship_type}")


        # Check for the sheet in both workbooks
        def get_sheet(workbook_hierarchy, workbook_operating, sheet_name):
            try:
                # First, try to get the sheet from the hierarchy workbook
                return workbook_hierarchy[sheet_name]
            except KeyError:
                # If not found, try to get it from the operating parameters workbook
                try:
                    return workbook_operating[sheet_name]
                except KeyError:
                    # If still not found, raise an error
                    raise KeyError(f"Sheet '{sheet_name}' does not exist in either workbook.")


        # Get the sheets from both workbooks
        try:
            sheet = get_sheet(wb_hierarchy, wb_operating_params, sheet_name)
            target_sheet = get_sheet(wb_hierarchy, wb_operating_params, target_sheet_name)
        except KeyError as e:
            print(f"Error: Sheet not found - {e}")
            continue

        # Get the header rows from the loaded sheets
        header_row = [cell.value for cell in sheet[1]]
        target_header_row = [cell.value for cell in target_sheet[1]]

        # Get the index of the join columns
        try:
            join_column_index = header_row.index(join_column)
        except ValueError:
            print(f"Error: Join column '{join_column}' not found in sheet '{sheet_name}'.")
            continue

        try:
            target_join_column_index = target_header_row.index(target_join_column)
        except ValueError:
            print(f"Error: Target join column '{target_join_column}' not found in sheet '{target_sheet_name}'.")
            continue

        # Collect data to create relationships
        relationship_data = []

        # Iterate through rows to find matching values in both sheets
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            join_column_value = row[join_column_index]

            # Find the matching row in the target sheet
            for target_row in target_sheet.iter_rows(min_row=2, values_only=True):
                if target_row[target_join_column_index] == join_column_value:
                    relationship_data.append({
                        "query": f"""
                            MATCH (a:{sheet_name} {{ {join_column}: $join_column_value }})
                            MATCH (b:{target_sheet_name} {{ {target_join_column}: $join_column_value }})
                            MERGE (a)-[:{relationship_type}]->(b)
                        """,
                        "parameters": {
                            "join_column_value": join_column_value
                        }
                    })

        # Execute relationship creation queries
        for rel_query in relationship_data:
            try:
                graph.run(rel_query["query"], **rel_query["parameters"])
                print(f"Relationship {relationship_type} created between {sheet_name} and {target_sheet_name}.")
            except Exception as e:
                print(f"Error creating relationship: {e}")